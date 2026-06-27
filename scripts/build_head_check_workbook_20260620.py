#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build a reviewer-friendly Excel workbook from the TikTok head-check list.

Scope: the 378 "collision-suspect" head videos (passes_china_signal_filter == False)
— these are the contested ones that decide the reach band (撞词噪声 vs 本土化出海).

Sheets:
- 填写说明  : 口径 + 三个下拉列怎么填 + §5.2 红线
- 核查清单  : 378 行, manual columns have dropdown validation, url hyperlinked
- 进度统计  : per-project COUNTIF/COUNTIFS formulas (待核查/已填/本土化/撞词)

Input : data/derived/tiktok_head_check_list_20260620.csv
Output: data/review/tiktok_head_check_priority378_20260620.xlsx
"""
from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "data/derived/tiktok_head_check_list_20260620.csv"
OUT = ROOT / "data/review/tiktok_head_check_priority378_20260620.xlsx"

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="16557A")
HEAD_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
INPUT_FILL = PatternFill("solid", fgColor="FFF6D5")   # 人工填写列底色(待填)
BASE_FONT = Font(name=FONT, size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

REL_OPTS = "相关,不相关,拿不准"
NAT_OPTS = "撞词无关,有中国语境,本土化无中国标记"
FORM_OPTS = "教学,展示表演,视觉奇观,vlog记录,卖货,其他"

# 核查清单列定义: (表头, 源字段, 列宽, 数字格式)
COLS = [
    ("编号", "check_id", 11, None),
    ("项目", "project_name", 16, None),
    ("播放量", "play_count", 11, "#,##0"),
    ("累计占比", "cum_play_share", 9, "0.0%"),
    ("命中搜索词", "source_term", 12, None),
    ("作者", "author_handle", 14, None),
    ("链接(点开看视频)", "url", 26, None),
    ("文案", "caption", 46, None),
    ("标签", "hashtags", 26, None),
    ("是否相关 ▼", None, 11, None),
    ("性质 ▼", None, 16, None),
    ("内容形态 ▼", None, 12, None),
    ("填写人", None, 9, None),
    ("备注", None, 28, None),
]
MANUAL_START = 10   # 第10列(J)起为人工填写列


def load_suspect() -> list[dict[str, str]]:
    rows = [r for r in csv.DictReader(SRC.open(encoding="utf-8-sig"))
            if r["passes_china_signal_filter"] == "False"]
    rows.sort(key=lambda r: (r["project_name"], -int(r["play_count"])))
    return rows


def build_instructions(ws, batch_label: str = "") -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 110
    title = "中国非遗 TikTok 头部视频 · 人工核查填写说明"
    if batch_label:
        title += f"（{batch_label}）"
    blocks = [
        (title, "title"),
        ("", None),
        ("为什么要核查这批视频", "h"),
        ("机器判为 likely(可能相关)、但视频本身没有任何中国文本信号(标题/文案无中文、无 China/Chinese)。", None),
        ("这类视频有两种完全相反的可能,必须人工看视频才能区分:", None),
        ("  ① 撞词噪声——同名异物,跟中国非遗无关(该排除);", None),
        ("  ② 本土化出海——外国人真在做这件中国非遗,只是用英文、不打中国标签(这是真触达,且是出海成功的关键证据,绝不能当噪声删!)。", None),
        ("本表是各项目头部高播放视频(累计覆盖约 80% 播放量)中的撞词嫌疑项,共 378 条,优先核查。", None),
        ("", None),
        ("三个下拉列怎么填(都在「核查清单」表,黄色列)", "h"),
        ("1) 是否相关:视频是否在传播「中国被列入该 UNESCO 项目」的那个对象。", None),
        ("     相关 / 不相关 / 拿不准(信息不足、画面或语言判断不了时填拿不准,不要硬猜)。", None),
        ("2) 性质:", None),
        ("     撞词无关——同名异物。例:taichi 撞冲绳地名/游戏、papercut 撞「割伤(paper cut)」、calligraphy 撞各国手写。", None),
        ("     有中国语境——相关,且有中国/华人/中文线索。", None),
        ("     本土化无中国标记——外国人真做这件中国非遗,但全英文、无中文、不打 #china。例:西方博主英文教太极、老外剪窗花。", None),
        ("3) 内容形态:教学 / 展示表演 / 视觉奇观 / vlog记录 / 卖货 / 其他。", None),
        ("", None),
        ("口径红线(摘自规格 §5.2,务必遵守)", "h"),
        ("• 只算中国的:蒙古国呼麦、日本书道、墨西哥剪纸 = 不相关;分不清国别 = 拿不准。", None),
        ("• 粤剧只算广东粤剧(Cantonese opera),不算浙江越剧;分不清 = 拿不准。", None),
        ("• 台湾布袋戏算福建木偶戏(分支流变,宽口径纳入)。", None),
        ("• 传播技艺/艺术/习俗才算;纯卖成品、现代工业品、奶茶/现代茶饮 = 不算。", None),
        ("• 分不清的一律填「拿不准」,不要猜——猜错会污染后面所有结论。", None),
        ("", None),
        ("填完后", "h"),
        ("每行填上「填写人」;有疑问写在「备注」。填好整表发回即可,我们会按「人工 > 机器」覆盖重算触达。", None),
    ]
    r = 1
    for text, kind in blocks:
        c = ws.cell(row=r, column=2, value=text)
        if kind == "title":
            c.font = Font(name=FONT, bold=True, size=15, color="16557A")
        elif kind == "h":
            c.font = Font(name=FONT, bold=True, size=11, color="16557A")
        else:
            c.font = Font(name=FONT, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1


def build_checklist(ws, rows: list[dict[str, str]]) -> int:
    # header
    for ci, (head, *_rest) in enumerate(COLS, 1):
        c = ws.cell(row=1, column=ci, value=head)
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = COLS[ci - 1][2]
    # rows
    for ri, row in enumerate(rows, start=2):
        for ci, (head, field, _w, numfmt) in enumerate(COLS, 1):
            cell = ws.cell(row=ri, column=ci)
            cell.font = BASE_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top",
                                       wrap_text=(field in ("caption", "hashtags") or ci >= MANUAL_START))
            if field is None:
                cell.fill = INPUT_FILL       # 人工填写列
                continue
            val = row.get(field, "")
            if field == "play_count":
                cell.value = int(val or 0)
            elif field == "cum_play_share":
                cell.value = float(val or 0)
            elif field == "url":
                cell.value = val
                if val:
                    cell.hyperlink = val
                    cell.font = Font(name=FONT, size=10, color="0563C1", underline="single")
            else:
                cell.value = val
            if numfmt:
                cell.number_format = numfmt
    last = len(rows) + 1
    # dropdowns
    for col_idx, opts in ((MANUAL_START, REL_OPTS), (MANUAL_START + 1, NAT_OPTS), (MANUAL_START + 2, FORM_OPTS)):
        dv = DataValidation(type="list", formula1=f'"{opts}"', allow_blank=True, showErrorMessage=True)
        dv.error = "请从下拉选项中选择"
        dv.errorTitle = "无效输入"
        ws.add_data_validation(dv)
        letter = get_column_letter(col_idx)
        dv.add(f"{letter}2:{letter}{last}")
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{last}"
    ws.row_dimensions[1].height = 30
    return last


def build_progress(ws, projects: list[str], list_name: str, last_row: int) -> None:
    ws.sheet_view.showGridLines = False
    headers = ["项目", "待核查数", "已填", "撞词无关", "有中国语境", "本土化无中国标记", "拿不准/相关存疑"]
    widths = [16, 10, 8, 10, 11, 16, 14]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = HEAD_FILL; c.font = HEAD_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    proj_rng = f"'{list_name}'!$B$2:$B${last_row}"
    rel_rng = f"'{list_name}'!$J$2:$J${last_row}"
    nat_rng = f"'{list_name}'!$K$2:$K${last_row}"
    for ri, proj in enumerate(projects, start=2):
        pcell = f"$A{ri}"
        ws.cell(row=ri, column=1, value=proj).font = BASE_FONT
        ws.cell(row=ri, column=2, value=f'=COUNTIF({proj_rng},{pcell})').font = BASE_FONT
        ws.cell(row=ri, column=3, value=f'=COUNTIFS({proj_rng},{pcell},{rel_rng},"<>")').font = BASE_FONT
        ws.cell(row=ri, column=4, value=f'=COUNTIFS({proj_rng},{pcell},{nat_rng},"撞词无关")').font = BASE_FONT
        ws.cell(row=ri, column=5, value=f'=COUNTIFS({proj_rng},{pcell},{nat_rng},"有中国语境")').font = BASE_FONT
        ws.cell(row=ri, column=6, value=f'=COUNTIFS({proj_rng},{pcell},{nat_rng},"本土化无中国标记")').font = BASE_FONT
        ws.cell(row=ri, column=7, value=f'=COUNTIFS({proj_rng},{pcell},{rel_rng},"拿不准")').font = BASE_FONT
    tr = len(projects) + 2
    ws.cell(row=tr, column=1, value="合计").font = Font(name=FONT, bold=True, size=10)
    for col in range(2, 8):
        L = get_column_letter(col)
        c = ws.cell(row=tr, column=col, value=f"=SUM({L}2:{L}{tr-1})")
        c.font = Font(name=FONT, bold=True, size=10)
    ws.freeze_panes = "A2"


def partition_projects(rows: list[dict[str, str]], n: int = 3) -> list[list[str]]:
    """整项目不拆分;贪心装箱使各批总条数尽量均衡(大项目先放进当前最小的批)。"""
    counts: dict[str, int] = {}
    for r in rows:
        counts[r["project_name"]] = counts.get(r["project_name"], 0) + 1
    bins: list[list[str]] = [[] for _ in range(n)]
    loads = [0] * n
    for proj, cnt in sorted(counts.items(), key=lambda kv: (-kv[1], kv[0])):
        i = loads.index(min(loads))
        bins[i].append(proj)
        loads[i] += cnt
    return bins


def build_batch(rows: list[dict[str, str]], projects: list[str], label: str, out: Path) -> int:
    sub = [r for r in rows if r["project_name"] in set(projects)]
    sub.sort(key=lambda r: (r["project_name"], -int(r["play_count"])))
    projects_sorted = sorted(projects)
    wb = Workbook()
    ws_info = wb.active
    ws_info.title = "填写说明"
    build_instructions(ws_info, label)
    ws_list = wb.create_sheet("核查清单")
    last = build_checklist(ws_list, sub)
    ws_prog = wb.create_sheet("进度统计")
    build_progress(ws_prog, projects_sorted, "核查清单", last)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return len(sub)


def main() -> int:
    rows = load_suspect()

    # 单一全量工作簿(留存)
    all_projects = sorted({r["project_name"] for r in rows})
    wb = Workbook()
    ws_info = wb.active
    ws_info.title = "填写说明"
    build_instructions(ws_info)
    ws_list = wb.create_sheet("核查清单")
    last = build_checklist(ws_list, rows)
    ws_prog = wb.create_sheet("进度统计")
    build_progress(ws_prog, all_projects, "核查清单", last)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT.name} | rows={len(rows)} projects={len(all_projects)}")

    # 按 A/B/C 三批拆分(整项目不拆,条数均衡),各出独立文件
    bins = partition_projects(rows, 3)
    for label, projects in zip(["A批", "B批", "C批"], bins):
        out = OUT.parent / f"tiktok_head_check_{label[0]}_20260620.xlsx"
        n = build_batch(rows, projects, label, out)
        print(f"wrote {out.name} | {label} | rows={n} projects={len(projects)} -> {sorted(projects)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
